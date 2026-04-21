#!/usr/bin/env python3
"""Inventory Q1 2026 Daylily omics analysis S3 exports.

This script is intentionally read-only against S3. It only lists objects and
streams small metadata/config files needed for reporting.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import re
import shlex
from collections import Counter
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import boto3
from botocore.exceptions import ClientError

BUCKET = "lsmc-dayoa-omics-analysis-us-west-2"
EXPORTS = [
    "FSxLustre20260122T043503Z",
    "FSxLustre20260122T112533Z",
    "FSxLustre20260122T113142Z",
    "FSxLustre20260211T122354Z",
    "FSxLustre20260213T142513Z",
    "FSxLustre20260216T130001Z",
    "FSxLustre20260309T122755Z",
]
OUTPUT_COLUMNS = [
    "fsx_export",
    "export_month",
    "analysis_code",
    "repo_s3_uri",
    "status",
    "has_day_cmd_log",
    "has_success_marker",
    "hg38_result_prefixes",
    "hg38_result_object_count",
    "sample_count",
    "unit_count",
    "sample_ids",
    "run_ids",
    "platforms",
    "coverage_or_experiment_ids",
    "first_command_time",
    "last_command_time",
    "command_count",
    "dryrun_command_count",
    "pipeline_command",
    "pipeline_targets",
    "pipeline_genome",
    "pipeline_profile",
    "pipeline_jobs",
    "total_runtime_seconds",
    "rule_count",
    "top_rules",
    "notes",
]
SMK_RE = re.compile(
    r"SMK>\s+D:(?P<time>\S+).*?\s+PWD:(?P<pwd>.*?)\s+/ CMD:\s+\((?P<cmd>.*)\)\s*$"
)
MAX_METADATA_BYTES = 8 * 1024 * 1024


@dataclass
class CommandEntry:
    fsx_export: str
    analysis_code: str
    command_time: str
    pwd: str
    command: str
    is_dryrun: bool


@dataclass
class InventoryRow:
    fsx_export: str
    export_month: str
    analysis_code: str
    repo_s3_uri: str
    status: str = "metadata_missing"
    has_day_cmd_log: bool = False
    has_success_marker: bool = False
    hg38_result_prefixes: str = ""
    hg38_result_object_count: int = 0
    sample_count: int = 0
    unit_count: int = 0
    sample_ids: str = ""
    run_ids: str = ""
    platforms: str = ""
    coverage_or_experiment_ids: str = ""
    first_command_time: str = ""
    last_command_time: str = ""
    command_count: int = 0
    dryrun_command_count: int = 0
    pipeline_command: str = ""
    pipeline_targets: str = ""
    pipeline_genome: str = ""
    pipeline_profile: str = ""
    pipeline_jobs: str = ""
    total_runtime_seconds: str = ""
    rule_count: int = 0
    top_rules: str = ""
    notes: str = ""
    note_items: list[str] = field(default_factory=list, repr=False)

    def finalize(self) -> dict[str, Any]:
        payload = asdict(self)
        payload.pop("note_items", None)
        payload["notes"] = "; ".join(dict.fromkeys(self.note_items)) or "none"
        return payload


class S3Inventory:
    def __init__(self, profile: str, region: str) -> None:
        session = boto3.Session(profile_name=profile, region_name=region)
        self.s3 = session.client("s3")
        self.profile = profile
        self.region = region

    def list_common_prefixes(self, prefix: str, delimiter: str = "/") -> list[str]:
        paginator = self.s3.get_paginator("list_objects_v2")
        prefixes: list[str] = []
        for page in paginator.paginate(Bucket=BUCKET, Prefix=prefix, Delimiter=delimiter):
            for item in page.get("CommonPrefixes", []):
                value = item.get("Prefix")
                if value:
                    prefixes.append(str(value))
        return sorted(set(prefixes))

    def list_keys(self, prefix: str, *, max_keys: int | None = None) -> list[str]:
        paginator = self.s3.get_paginator("list_objects_v2")
        keys: list[str] = []
        for page in paginator.paginate(Bucket=BUCKET, Prefix=prefix):
            for item in page.get("Contents", []):
                key = item.get("Key")
                if key:
                    keys.append(str(key))
                    if max_keys is not None and len(keys) >= max_keys:
                        return keys
        return keys

    def count_objects(self, prefix: str) -> int:
        paginator = self.s3.get_paginator("list_objects_v2")
        count = 0
        for page in paginator.paginate(Bucket=BUCKET, Prefix=prefix):
            count += len(page.get("Contents", []))
        return count

    def object_exists(self, key: str) -> bool:
        try:
            self.s3.head_object(Bucket=BUCKET, Key=key)
        except ClientError as exc:
            code = exc.response.get("Error", {}).get("Code")
            if code in {"404", "NoSuchKey", "NotFound"}:
                return False
            raise
        return True

    def read_text(self, key: str) -> tuple[str | None, str | None]:
        try:
            head = self.s3.head_object(Bucket=BUCKET, Key=key)
        except ClientError as exc:
            code = exc.response.get("Error", {}).get("Code")
            if code in {"404", "NoSuchKey", "NotFound"}:
                return None, "missing"
            raise

        size = int(head.get("ContentLength") or 0)
        kwargs: dict[str, Any] = {"Bucket": BUCKET, "Key": key}
        note = None
        if size > MAX_METADATA_BYTES:
            kwargs["Range"] = f"bytes={max(size - MAX_METADATA_BYTES, 0)}-{size - 1}"
            note = f"truncated {key.rsplit('/', 1)[-1]} to last {MAX_METADATA_BYTES} bytes"
        body = self.s3.get_object(**kwargs)["Body"].read()
        return body.decode("utf-8", errors="replace"), note


def export_month(fsx_export: str) -> str:
    match = re.search(r"FSxLustre(\d{4})(\d{2})", fsx_export)
    if not match:
        return ""
    return f"{match.group(1)}-{match.group(2)}"


def unique_join(values: Iterable[str]) -> str:
    cleaned = sorted({str(value).strip() for value in values if str(value).strip()})
    return ",".join(cleaned)


def parse_tsv(text: str | None) -> list[dict[str, str]]:
    if not text:
        return []
    reader = csv.DictReader(io.StringIO(text), delimiter="\t")
    return [
        {str(key or "").strip(): str(value or "").strip() for key, value in row.items()}
        for row in reader
    ]


def first_present(row: dict[str, str], names: Iterable[str]) -> str:
    lowered = {key.lower(): value for key, value in row.items()}
    for name in names:
        value = lowered.get(name.lower())
        if value:
            return value
    return ""


def extract_coverage_values(rows: list[dict[str, str]]) -> list[str]:
    values: list[str] = []
    path_columns = [name for name in rows[0].keys() if name.upper().endswith("_PATH")] if rows else []
    for row in rows:
        experiment = first_present(row, ["EXPERIMENTID", "EXPERIMENT_ID", "experiment_id"])
        if experiment:
            values.append(experiment)
        for column in path_columns:
            values.extend(re.findall(r"(?<![\w.])\d+(?:\.\d+)?x(?!\w)", row.get(column, "")))
    return values


def parse_commands(
    text: str | None,
    *,
    fsx_export: str,
    analysis_code: str,
) -> list[CommandEntry]:
    entries: list[CommandEntry] = []
    if not text:
        return entries
    for line in text.splitlines():
        match = SMK_RE.search(line)
        if not match:
            continue
        command = match.group("cmd").strip()
        entries.append(
            CommandEntry(
                fsx_export=fsx_export,
                analysis_code=analysis_code,
                command_time=match.group("time"),
                pwd=match.group("pwd").strip(),
                command=command,
                is_dryrun=is_dryrun_command(command),
            )
        )
    return entries


def is_dryrun_command(command: str) -> bool:
    try:
        parts = shlex.split(command)
    except ValueError:
        parts = command.split()
    return "-n" in parts or "--dry-run" in parts or "dryrun" in command.lower()


def selected_pipeline_command(entries: list[CommandEntry]) -> str:
    non_dryrun = [entry for entry in entries if not entry.is_dryrun]
    source = non_dryrun or entries
    return source[-1].command if source else ""


def option_value(parts: list[str], names: set[str]) -> str:
    for index, token in enumerate(parts):
        for name in names:
            if token == name and index + 1 < len(parts):
                return parts[index + 1]
            if token.startswith(f"{name}="):
                return token.split("=", 1)[1]
    return ""


def extract_pipeline_fields(command: str) -> tuple[str, str, str, str]:
    if not command:
        return "", "", "", ""
    try:
        parts = shlex.split(command)
    except ValueError:
        parts = command.split()

    genome = ""
    for token in parts:
        if re.fullmatch(r"hg38[\w-]*", token):
            genome = token
            break
    if not genome:
        match = re.search(r"\bhg38[\w-]*\b", command)
        genome = match.group(0) if match else ""

    profile = option_value(parts, {"--profile", "--workflow-profile"})
    jobs = option_value(parts, {"--jobs", "-j", "--cores"})
    targets = extract_targets(parts)
    return targets, genome, profile, jobs


def extract_targets(parts: list[str]) -> str:
    option_tokens_with_values = {
        "--profile",
        "--workflow-profile",
        "--jobs",
        "-j",
        "--cores",
        "--configfile",
        "--directory",
        "--snakefile",
        "-s",
    }
    targets: list[str] = []
    skip_next = False
    for index, token in enumerate(parts):
        if skip_next:
            skip_next = False
            continue
        if index == 0 or token in {"dy-r", "dy-a", "snakemake", "slurm"}:
            continue
        if token in option_tokens_with_values:
            skip_next = True
            continue
        if token.startswith("-") or "=" in token:
            continue
        if re.fullmatch(r"hg38[\w-]*", token):
            continue
        if "/" in token and not token.startswith("results/"):
            continue
        targets.append(token)
    return ",".join(dict.fromkeys(targets))


def parse_stats(text: str | None) -> tuple[str, int, str, str | None]:
    if not text:
        return "", 0, "", "missing day_pipe_stats.json"
    try:
        payload = json.loads(text)
    except json.JSONDecodeError as exc:
        return "", 0, "", f"failed to parse day_pipe_stats.json: {exc}"
    runtime = payload.get("total_runtime_seconds", payload.get("total_runtime", ""))
    runtime_text = "" if runtime in (None, "") else str(round(float(runtime), 3))
    rules = payload.get("rules", {})
    if not isinstance(rules, dict):
        return runtime_text, 0, "", "day_pipe_stats.json rules is not a mapping"
    top = []
    for name, stats in rules.items():
        score = 0.0
        if isinstance(stats, dict):
            try:
                score = float(stats.get("max-runtime", stats.get("mean-runtime", 0)) or 0)
            except (TypeError, ValueError):
                score = 0.0
        top.append((str(name), score))
    top_rules = ",".join(name for name, _score in sorted(top, key=lambda item: item[1], reverse=True)[:10])
    return runtime_text, len(rules), top_rules, None


def discover_repo_prefixes(client: S3Inventory) -> list[tuple[str, str, str]]:
    repos: list[tuple[str, str, str]] = []
    for fsx_export in EXPORTS:
        base = f"{fsx_export}/analysis_results/ubuntu/"
        for analysis_prefix in client.list_common_prefixes(base):
            analysis_code = analysis_prefix.rstrip("/").rsplit("/", 1)[-1]
            for child_prefix in client.list_common_prefixes(analysis_prefix):
                if child_prefix.endswith("/daylily-omics-analysis/"):
                    repos.append((fsx_export, analysis_code, child_prefix))
    return sorted(set(repos))


def hg38_prefixes_and_count(client: S3Inventory, repo_prefix: str) -> tuple[list[str], int]:
    day_prefix = f"{repo_prefix}results/day/"
    common = client.list_common_prefixes(day_prefix)
    hg38_prefixes = [
        prefix
        for prefix in common
        if prefix.rstrip("/").rsplit("/", 1)[-1] == "hg38"
        or prefix.rstrip("/").rsplit("/", 1)[-1].startswith("hg38_")
    ]
    if not hg38_prefixes:
        for key in client.list_keys(day_prefix, max_keys=5000):
            rest = key[len(day_prefix) :]
            first = rest.split("/", 1)[0]
            if first == "hg38" or first.startswith("hg38_"):
                hg38_prefixes.append(f"{day_prefix}{first}/")
        hg38_prefixes = sorted(set(hg38_prefixes))
    count = sum(client.count_objects(prefix) for prefix in hg38_prefixes)
    return hg38_prefixes, count


def classify_status(
    *,
    has_success_marker: bool,
    hg38_count: int,
    metadata_missing: bool,
) -> str:
    if metadata_missing:
        return "metadata_missing"
    if has_success_marker and hg38_count > 0:
        return "complete"
    if hg38_count > 0:
        return "partial_results"
    return "no_hg38_results"


def inventory_repo(
    client: S3Inventory,
    *,
    fsx_export: str,
    analysis_code: str,
    repo_prefix: str,
) -> tuple[InventoryRow, list[CommandEntry]]:
    row = InventoryRow(
        fsx_export=fsx_export,
        export_month=export_month(fsx_export),
        analysis_code=analysis_code,
        repo_s3_uri=f"s3://{BUCKET}/{repo_prefix}",
    )

    hg38_prefixes, hg38_count = hg38_prefixes_and_count(client, repo_prefix)
    row.hg38_result_prefixes = ",".join(f"s3://{BUCKET}/{prefix}" for prefix in hg38_prefixes)
    row.hg38_result_object_count = hg38_count
    row.has_success_marker = client.object_exists(f"{repo_prefix}daylily.successful_run")

    day_cmd_text, note = client.read_text(f"{repo_prefix}day_cmd.log")
    if note and note != "missing":
        row.note_items.append(note)
    row.has_day_cmd_log = day_cmd_text is not None
    commands = parse_commands(day_cmd_text, fsx_export=fsx_export, analysis_code=analysis_code)
    row.command_count = len(commands)
    row.dryrun_command_count = sum(1 for entry in commands if entry.is_dryrun)
    if commands:
        row.first_command_time = commands[0].command_time
        row.last_command_time = commands[-1].command_time
    selected = selected_pipeline_command(commands)
    row.pipeline_command = selected
    (
        row.pipeline_targets,
        row.pipeline_genome,
        row.pipeline_profile,
        row.pipeline_jobs,
    ) = extract_pipeline_fields(selected)

    samples_text, samples_note = client.read_text(f"{repo_prefix}config/samples.tsv")
    units_text, units_note = client.read_text(f"{repo_prefix}config/units.tsv")
    for metadata_note in (samples_note, units_note):
        if metadata_note and metadata_note != "missing":
            row.note_items.append(metadata_note)
    sample_rows = parse_tsv(samples_text)
    unit_rows = parse_tsv(units_text)
    sample_ids = [
        first_present(item, ["SAMPLEID", "SAMPLE_ID", "sample_id"])
        for item in sample_rows
    ]
    row.sample_ids = unique_join(sample_ids)
    row.sample_count = len({value for value in sample_ids if value})
    row.unit_count = len(unit_rows)
    row.run_ids = unique_join(
        first_present(item, ["RUNID", "RUN_ID", "run_id"]) for item in unit_rows
    )
    row.platforms = unique_join(
        [
            *(
                first_present(item, ["SEQ_VENDOR", "vendor", "sequencer"])
                for item in unit_rows
            ),
            *(
                first_present(item, ["SEQ_PLATFORM", "platform"])
                for item in unit_rows
            ),
        ]
    )
    row.coverage_or_experiment_ids = unique_join(extract_coverage_values(unit_rows))

    stats_text, stats_note = client.read_text(f"{repo_prefix}day_pipe_stats.json")
    if stats_note and stats_note != "missing":
        row.note_items.append(stats_note)
    (
        row.total_runtime_seconds,
        row.rule_count,
        row.top_rules,
        stats_parse_note,
    ) = parse_stats(stats_text)
    if stats_parse_note and stats_parse_note != "missing day_pipe_stats.json":
        row.note_items.append(stats_parse_note)

    missing_metadata = []
    if not row.has_day_cmd_log:
        missing_metadata.append("day_cmd.log")
    if samples_text is None:
        missing_metadata.append("config/samples.tsv")
    if units_text is None:
        missing_metadata.append("config/units.tsv")
    if missing_metadata:
        row.note_items.append("missing metadata: " + ",".join(missing_metadata))

    row.status = classify_status(
        has_success_marker=row.has_success_marker,
        hg38_count=row.hg38_result_object_count,
        metadata_missing=bool(missing_metadata),
    )
    return row, commands


def write_tsv(rows: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=OUTPUT_COLUMNS,
            delimiter="\t",
            lineterminator="\n",
        )
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in OUTPUT_COLUMNS})


def write_json(payload: dict[str, Any], output_path: Path) -> None:
    output_path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def write_workbook(
    rows: list[dict[str, Any]],
    commands: list[CommandEntry],
    output_path: Path,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    status_counts = Counter(row["status"] for row in rows)
    ws.append(["Metric", "Value"])
    ws.append(["Total analysis dirs", len(rows)])
    for status in ["complete", "partial_results", "no_hg38_results", "metadata_missing"]:
        ws.append([status, status_counts.get(status, 0)])
    ws.append(["Completed hg38 object count", sum(int(row["hg38_result_object_count"]) for row in rows if row["status"] == "complete")])
    style_sheet(ws, table_name="SummaryTable")

    write_sheet(wb, "All dirs", OUTPUT_COLUMNS, rows)
    write_sheet(
        wb,
        "Completed dirs",
        OUTPUT_COLUMNS,
        [row for row in rows if row["status"] == "complete"],
    )
    write_sheet(
        wb,
        "Incomplete dirs",
        OUTPUT_COLUMNS,
        [row for row in rows if row["status"] != "complete"],
    )
    command_columns = ["fsx_export", "analysis_code", "command_time", "pwd", "is_dryrun", "command"]
    command_rows = [
        {
            "fsx_export": entry.fsx_export,
            "analysis_code": entry.analysis_code,
            "command_time": entry.command_time,
            "pwd": entry.pwd,
            "is_dryrun": entry.is_dryrun,
            "command": entry.command,
        }
        for entry in commands
    ]
    write_sheet(wb, "Command logs", command_columns, command_rows)

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for column_cells in sheet.columns:
            values = [str(cell.value or "") for cell in column_cells[:200]]
            width = min(max(max((len(value) for value in values), default=8) + 2, 10), 70)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    wb.save(output_path)


def style_sheet(ws: Any, *, table_name: str) -> None:
    from openpyxl.worksheet.table import Table, TableStyleInfo

    if ws.max_row < 2 or ws.max_column < 1:
        return
    table = Table(displayName=table_name, ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def safe_table_name(sheet_name: str) -> str:
    return re.sub(r"\W+", "", sheet_name)[:20] + "Table"


def write_sheet(
    wb: Any,
    sheet_name: str,
    columns: list[str],
    rows: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column, "") for column in columns])
    style_sheet(ws, table_name=safe_table_name(sheet_name))


def write_runbook(rows: list[dict[str, Any]], commands: list[CommandEntry], output_path: Path) -> None:
    examples = command_examples(commands)
    status_counts = Counter(row["status"] for row in rows)
    lines = [
        "# dy-r Runbook For Q1 2026 Daylily Omics Analysis Re-Runs",
        "",
        "This runbook is generated from observed Q1 2026 exported Daylily analysis repositories.",
        "",
        "## Inventory Summary",
        "",
        f"- Total analysis dirs: {len(rows)}",
        f"- Complete: {status_counts.get('complete', 0)}",
        f"- Partial results: {status_counts.get('partial_results', 0)}",
        f"- No hg38 results: {status_counts.get('no_hg38_results', 0)}",
        f"- Metadata missing: {status_counts.get('metadata_missing', 0)}",
        "",
        "## Standard Headnode Session",
        "",
        "Use the Daylily headnode shell configured for `ubuntu` in a bash login shell.",
        "",
        "```bash",
        "cd /fsx/analysis_results/ubuntu/<analysis-code>/daylily-omics-analysis",
        "source ~/.bashrc",
        "command -v dy-r",
        "command -v dy-a",
        "```",
        "",
        "## Recreate Config Files",
        "",
        "```bash",
        "mkdir -p config",
        "# Copy or regenerate the staged manifests:",
        "cp /fsx/data/staged_sample_data/<stage-dir>/*_samples.tsv config/samples.tsv",
        "cp /fsx/data/staged_sample_data/<stage-dir>/*_units.tsv config/units.tsv",
        "",
        "head -n 3 config/samples.tsv",
        "head -n 3 config/units.tsv",
        "python - <<'PY'",
        "import csv",
        "for path in ['config/samples.tsv', 'config/units.tsv']:",
        "    with open(path, newline='') as handle:",
        "        rows = list(csv.DictReader(handle, delimiter='\\t'))",
        "    print(path, len(rows), rows[0].keys() if rows else 'EMPTY')",
        "PY",
        "```",
        "",
        "## Launch Pattern",
        "",
        "Use the observed profile convention, then run a dry-run before the real run.",
        "",
        "```bash",
        "dy-a slurm hg38",
        "dy-r <target> -n",
        "dy-r <target>",
        "",
        "# Broad-profile runs observed in this inventory use:",
        "dy-a slurm hg38_broad",
        "dy-r <target> -n",
        "dy-r <target>",
        "```",
        "",
        "## Observed Command Examples",
        "",
    ]
    if examples:
        for label, command in examples:
            lines.extend([f"### {label}", "", "```bash", command, "```", ""])
    else:
        lines.extend(["No `SMK>` command entries were found in the scanned metadata.", ""])

    lines.extend(
        [
            "## Validation Checklist",
            "",
            "- Confirm `config/samples.tsv` has the expected `SAMPLEID` values.",
            "- Confirm `config/units.tsv` has expected `RUNID`, `SAMPLEID`, platform, and path columns.",
            "- Run the dry-run command first and inspect missing input messages.",
            "- Launch the real command only after the dry-run resolves.",
            "- Check `day_cmd.log`, `day_pipe_stats.json`, and `daylily.successful_run` after completion.",
        ]
    )
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def command_examples(commands: list[CommandEntry]) -> list[tuple[str, str]]:
    selected: list[tuple[str, str]] = []
    wanted = {
        "ILMN": re.compile(r"ILMN|NOVASEQ|bwa|deep|snakemake|dy-r", re.I),
        "ONT": re.compile(r"ONT|nanopore", re.I),
        "PacBio": re.compile(r"PACBIO|PB_|hifi|pacbio", re.I),
        "Ultima": re.compile(r"ULTIMA|UG_", re.I),
        "Roche": re.compile(r"ROCHE|AVITI|kapa", re.I),
        "Hybrid": re.compile(r"hybrid|ONT|PACBIO|ULTIMA", re.I),
    }
    seen_commands: set[str] = set()
    for label, pattern in wanted.items():
        for entry in reversed(commands):
            if entry.command in seen_commands:
                continue
            if pattern.search(entry.command):
                selected.append((label, entry.command))
                seen_commands.add(entry.command)
                break
    return selected


def validate_outputs(rows: list[dict[str, Any]], tsv_path: Path, xlsx_path: Path) -> list[str]:
    issues: list[str] = []
    lines = tsv_path.read_text(encoding="utf-8").splitlines()
    expected_tabs = len(OUTPUT_COLUMNS) - 1
    if not lines or lines[0].split("\t") != OUTPUT_COLUMNS:
        issues.append("TSV header does not match expected columns")
    for index, line in enumerate(lines[1:], start=2):
        if line.count("\t") != expected_tabs:
            issues.append(f"TSV row {index} has unstable tab count")
    for row in rows:
        if row["status"] == "complete" and int(row["hg38_result_object_count"]) <= 0:
            issues.append(f"complete row has zero hg38 objects: {row['repo_s3_uri']}")
        if "202604" in row["fsx_export"]:
            issues.append("April export included unexpectedly")

    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True)
    expected_sheets = {"Summary", "All dirs", "Completed dirs", "Incomplete dirs", "Command logs"}
    if set(wb.sheetnames) != expected_sheets:
        issues.append(f"Workbook sheets mismatch: {wb.sheetnames}")
    all_dirs = wb["All dirs"]
    if all_dirs.max_row != len(rows) + 1:
        issues.append("Workbook All dirs row count does not match TSV")
    return issues


def run(args: argparse.Namespace) -> int:
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    client = S3Inventory(profile=args.profile, region=args.region)
    repos = discover_repo_prefixes(client)

    rows: list[InventoryRow] = []
    commands: list[CommandEntry] = []
    for fsx_export, analysis_code, repo_prefix in repos:
        row, command_entries = inventory_repo(
            client,
            fsx_export=fsx_export,
            analysis_code=analysis_code,
            repo_prefix=repo_prefix,
        )
        rows.append(row)
        commands.extend(command_entries)

    row_payloads = [
        row.finalize()
        for row in sorted(rows, key=lambda item: (item.fsx_export, item.analysis_code, item.repo_s3_uri))
    ]
    commands = sorted(commands, key=lambda item: (item.fsx_export, item.analysis_code, item.command_time, item.command))

    tsv_path = output_dir / "analysis_dirs.tsv"
    xlsx_path = output_dir / "analysis_dirs.xlsx"
    manifest_path = output_dir / "inventory_manifest.json"
    runbook_path = output_dir / "dy-r-runbook.md"

    write_tsv(row_payloads, tsv_path)
    write_workbook(row_payloads, commands, xlsx_path)
    write_runbook(row_payloads, commands, runbook_path)
    manifest = {
        "bucket": BUCKET,
        "profile": args.profile,
        "region": args.region,
        "scan_time": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "included_exports": EXPORTS,
        "excluded_exports": ["FSxLustre20260412T103705Z"],
        "repo_count": len(row_payloads),
        "command_count": len(commands),
        "object_count_method": "S3 ListObjectsV2 pagination by hg38 result prefix; no result payload downloads",
        "outputs": {
            "analysis_dirs_tsv": str(tsv_path),
            "analysis_dirs_xlsx": str(xlsx_path),
            "dy_r_runbook": str(runbook_path),
        },
    }
    issues = validate_outputs(row_payloads, tsv_path, xlsx_path)
    manifest["validation_issues"] = issues
    write_json(manifest, manifest_path)

    print(json.dumps({"output_dir": str(output_dir), "rows": len(row_payloads), "issues": issues}, indent=2))
    return 1 if issues else 0


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--profile", default="daylily-service-lsmc")
    parser.add_argument("--region", default="us-west-2")
    parser.add_argument(
        "--output-dir",
        default="reports/dayoa-q1-2026-analysis-inventory",
    )
    raise SystemExit(run(parser.parse_args()))


if __name__ == "__main__":
    main()
